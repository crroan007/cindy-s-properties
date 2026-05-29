"""
ETL: load all Airbnb monthly rental ledgers (Google Sheets exported to xlsx/)
into a single local SQLite DB: lauren_way_rental.db

Each monthly workbook has one tab per property (e.g. "1120 Lauren Way",
"260 East Taylors Crossing", ...) plus a "Totals" tab. Each property tab is a
small ledger: col A = line item, col B = Debits, col C = Credits, with a
Total/Net summary row near the bottom.

We populate:
  - raw_cells        : every non-empty cell of every file/tab (lossless capture)
  - ledger_entries   : structured line items (property, line_item, debit, credit)
  - monthly_totals   : per-property + grand-total Total Debits / Credits / Net
  - files            : one row per source workbook
"""
import openpyxl, glob, os, re, sqlite3, sys

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX_DIR = os.path.join(ROOT, "xlsx")
DB = os.path.join(ROOT, "lauren_way_rental.db")

# canonical property names keyed by a normalized lookup
def norm_key(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

CANON = {
    "260easttaylorscrossing": "260 East Taylors Crossing",
    "260easttaylorscrossing2": "260 East Taylors Crossing",
    "4645valaisct": "4645 Valais Ct",
    "4645valaiscourt": "4645 Valais Ct",
    "5525taylorroad": "5525 Taylor Road",
    "1120laurenway": "1120 Lauren Way",
    "115peachtreememorialdrive": "115 Peachtree Memorial Drive",
}

MONTHS = {
    1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
    7:"July",8:"August",9:"September",10:"October",11:"November",12:"December",
}

def canon_property(tab_name):
    k = norm_key(tab_name)
    if k in CANON:
        return CANON[k]
    # fuzzy: tab names sometimes have typos / interpolated digits / abbreviated
    low = str(tab_name).lower()
    if "valais" in low:                return "4645 Valais Ct"
    if "lauren" in low:                return "1120 Lauren Way"
    if "taylors crossing" in low or low.strip().startswith("260"): return "260 East Taylors Crossing"
    if "taylor road" in low or low.strip().startswith("5525"):     return "5525 Taylor Road"
    if "peachtree" in low or "memorial" in low or low.strip().startswith("115"): return "115 Peachtree Memorial Drive"
    return None  # not a recognized property tab

def is_totals_tab(tab_name):
    return norm_key(tab_name) == "totals"

def num(v):
    """Return float if cell is numeric, else None."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        t = v.strip().replace("$", "").replace(",", "")
        # handle accounting-style negatives or stray text
        if re.fullmatch(r"-?\d+(\.\d+)?", t):
            return float(t)
    return None

def cell_str(v):
    return "" if v is None else str(v).strip()

TOTAL_HDR = {"total", "total debits", "total credits", "net"}

def parse_period(name):
    """name like 2018-08, 2017-01b, 2016-totals."""
    m = re.fullmatch(r"(\d{4})-(\d{2})([a-z]?)", name)
    if m:
        return dict(kind="monthly", year=int(m.group(1)), month=int(m.group(2)),
                    variant=m.group(3) or "")
    m = re.fullmatch(r"(\d{4})-totals", name)
    if m:
        return dict(kind="year_totals", year=int(m.group(1)), month=None, variant="")
    return dict(kind="other", year=None, month=None, variant="")


def build():
    if os.path.exists(DB):
        os.remove(DB)
    con = sqlite3.connect(DB)
    c = con.cursor()
    c.executescript("""
    CREATE TABLE files(
        source_file TEXT PRIMARY KEY,
        kind TEXT, year INTEGER, month INTEGER, variant TEXT,
        n_tabs INTEGER, tab_names TEXT
    );
    CREATE TABLE raw_cells(
        source_file TEXT, tab TEXT, row INTEGER, col INTEGER,
        value TEXT
    );
    CREATE TABLE ledger_entries(
        source_file TEXT, year INTEGER, month INTEGER, period TEXT, variant TEXT,
        property TEXT, property_raw TEXT,
        line_item TEXT, debit REAL, credit REAL
    );
    CREATE TABLE monthly_totals(
        source_file TEXT, year INTEGER, month INTEGER, period TEXT, variant TEXT,
        property TEXT, property_raw TEXT,
        total_debits REAL, total_credits REAL, net REAL
    );
    """)

    files = sorted(glob.glob(os.path.join(XLSX_DIR, "*.xlsx")))
    n_entries = n_tot = 0
    for path in files:
        name = os.path.splitext(os.path.basename(path))[0]
        per = parse_period(name)
        period_label = name
        wb = openpyxl.load_workbook(path, data_only=True)
        c.execute("INSERT OR REPLACE INTO files VALUES(?,?,?,?,?,?,?)",
                  (name, per["kind"], per["year"], per["month"], per["variant"],
                   len(wb.sheetnames), "|".join(wb.sheetnames)))

        for ws in wb.worksheets:
            tab = ws.title
            rows = list(ws.iter_rows(values_only=True))
            # lossless raw capture
            for ri, row in enumerate(rows, start=1):
                for ci, val in enumerate(row, start=1):
                    s = cell_str(val)
                    if s != "":
                        c.execute("INSERT INTO raw_cells VALUES(?,?,?,?,?)",
                                  (name, tab, ri, ci, s))

            prop = canon_property(tab)
            totals_tab = is_totals_tab(tab)

            # structured parse only for monthly property tabs + monthly Totals tab
            if per["kind"] != "monthly":
                continue
            if prop is None and not totals_tab:
                continue

            # walk rows; capture line items + the Total/Net summary
            for ri, row in enumerate(rows):
                a = cell_str(row[0]) if len(row) > 0 else ""
                b = row[1] if len(row) > 1 else None
                cc = row[2] if len(row) > 2 else None
                d = row[3] if len(row) > 3 else None
                e = row[4] if len(row) > 4 else None

                la = a.lower()
                # summary marker row: col B/C say Total/Total Debits etc.
                if cell_str(b).lower() in TOTAL_HDR or cell_str(cc).lower() in TOTAL_HDR:
                    # numbers are on the NEXT row
                    if ri + 1 < len(rows):
                        nxt = rows[ri + 1]
                        td = num(nxt[1]) if len(nxt) > 1 else None
                        tcr = num(nxt[2]) if len(nxt) > 2 else None
                        net = None
                        # Net may sit in col D or E of the next row
                        for idx in (4, 3):
                            if len(nxt) > idx and num(nxt[idx]) is not None:
                                net = num(nxt[idx]); break
                        c.execute("INSERT INTO monthly_totals VALUES(?,?,?,?,?,?,?,?,?,?)",
                                  (name, per["year"], per["month"], period_label, per["variant"],
                                   prop if prop else "ALL", tab, td, tcr, net))
                        n_tot += 1
                    continue

                if totals_tab:
                    continue  # for Totals tab only keep the grand total summary

                # header row?
                if la in ("line item", "") and cell_str(b).lower() in ("debits", "credits", ""):
                    # skip header/blank-label rows unless they carry numbers
                    if la == "line item":
                        continue
                # a real line item: needs a text label and at least one numeric amount
                db = num(b); cr = num(cc)
                if a and a.lower() not in TOTAL_HDR and (db is not None or cr is not None):
                    c.execute("INSERT INTO ledger_entries VALUES(?,?,?,?,?,?,?,?,?,?)",
                              (name, per["year"], per["month"], period_label, per["variant"],
                               prop, tab, a, db, cr))
                    n_entries += 1

    con.commit()

    # report
    print("DB:", DB)
    for t in ("files", "raw_cells", "ledger_entries", "monthly_totals"):
        print(f"  {t}: {c.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0]} rows")
    print("\nFiles loaded:")
    for r in c.execute("SELECT source_file, kind, n_tabs, tab_names FROM files ORDER BY source_file"):
        print("  ", r[0], "|", r[1], "|", r[2], "tabs |", r[3])
    print("\nDistinct properties in ledger_entries:")
    for r in c.execute("SELECT property, COUNT(*) FROM ledger_entries GROUP BY property ORDER BY property"):
        print("  ", r)
    con.close()


if __name__ == "__main__":
    build()
