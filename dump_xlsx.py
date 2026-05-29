import openpyxl, glob, sys

for f in sorted(glob.glob('xlsx/*.xlsx')):
    print("\n############ FILE:", f, "############")
    wb = openpyxl.load_workbook(f, data_only=True)
    print("SHEETS:", wb.sheetnames)
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(c is not None and str(c).strip() != '' for c in row):
                rows.append([("" if c is None else c) for c in row])
        print(f"---- TAB {ws.title!r}  ({len(rows)} non-empty rows) ----")
        for r in rows:
            print(r)
